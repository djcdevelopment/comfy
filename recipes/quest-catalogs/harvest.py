#!/usr/bin/env python3
"""Harvest guild quest catalogs from their real sources.

This is the configurator seam of the absorption engine: sources.json says which guild,
which source, which adapter; each adapter's only contract is "emit quests that conform
to schema.md". Content passes through verbatim; anything odd lands in the anomalies
report for the guild to rule on — never silently fixed.

Usage:
  python harvest.py                     # harvest every enabled source in sources.json
  python harvest.py slayers-summons     # harvest one source by id

Outputs, per source:
  <output>.json            the canonical quest catalog
  <output>-anomalies.md    everything the guild should look at

The sheet-xlsx adapter needs openpyxl (pip install openpyxl). Everything else is
standard library only.
"""
import json
import os
import re
import sys

HERE = os.path.dirname(os.path.abspath(__file__))

SLOT_PATTERN = re.compile(r"(\w+):(?=\s|$)")
IMAGE_SLOT = re.compile(r"^image\d*$")
# the parameter that names which quest/badge the bot credits, and its value
NAME_PARAM = re.compile(r"(?:summons_type|badge_name):\s*(.*?)(?=\s+\w+:|$)")


# ---------------------------------------------------------------- shared helpers

def slugify(name):
    slug = re.sub(r"[^a-z0-9]+", "_", name.lower()).strip("_")
    return slug


def parse_evidence(bot_command, requirements_text):
    """Derive the evidence spec from the command's slots (machine truth) and the
    requirements text's emoji grammar (human truth). Mismatches are anomalies."""
    cmd = bot_command or ""
    slots = SLOT_PATTERN.findall(cmd)
    # the name parameter (summons_type:/badge_name:) is the selector, not evidence;
    # participants:/url slots may carry instruction text instead of being empty,
    # so those are detected by substring, not by the empty-slot pattern
    evidence_slots = [s for s in slots if s not in ("summons_type", "badge_name")]
    return {
        "screenshots": sum(1 for s in evidence_slots if IMAGE_SLOT.match(s)),
        "video_alternative": "\U0001F39E" in (requirements_text or ""),  # 🎞️
        "link": "summons_url:" in cmd,
        "group_turnin": "participants:" in cmd,
        "notes": "summons_notes:" in cmd or "badge_notes:" in cmd,
    }


def cross_check(quest, anomalies):
    """Compare emoji grammar in the text against the command slots."""
    text = quest["requirements"] or ""
    ev = quest["evidence"]
    camera_count = text.count("\U0001F4F8")  # 📸
    if quest["auto_checked"]:
        return
    if camera_count and ev["screenshots"] and camera_count != ev["screenshots"]:
        anomalies.append(
            f"**{quest['name']}**: requirements text shows {camera_count} camera emoji "
            f"but the bot command has {ev['screenshots']} image slot(s). "
            f"Which is right?"
        )
    if "\U0001F517" in text and not ev["link"]:  # 🔗
        anomalies.append(
            f"**{quest['name']}**: requirements mention a 🔗 link but the bot command "
            f"has no summons_url: slot."
        )
    if "\U0001F91C" in text and not ev["group_turnin"]:  # 🤜
        anomalies.append(
            f"**{quest['name']}**: requirements mention 🤜🤛 group turn-in but the bot "
            f"command has no participants: slot."
        )
    note = quest.get("evidence_note") or ""
    if ev["screenshots"] == 0 and re.search(r"screenshot|photo|image", note, re.I):
        anomalies.append(
            f"**{quest['name']}**: the evidence note asks for a screenshot/photo "
            f"({note[:60]!r}) but the bot command has no image slot."
        )


def finish_catalog(source, quests, anomalies):
    """Shared post-pass: ids unique, commands unique, wrap in the catalog envelope."""
    seen_ids = {}
    seen_commands = {}
    for q in quests:
        if q["quest_id"] in seen_ids:
            anomalies.append(
                f"**{q['name']}**: quest_id `{q['quest_id']}` collides with "
                f"**{seen_ids[q['quest_id']]}** — one of them needs a distinct name."
            )
        seen_ids[q["quest_id"]] = q["name"]

        cmd = q["bot_command"]
        if cmd:
            m = NAME_PARAM.search(cmd)
            credited = m.group(1).strip() if m else None
            if credited is None:
                anomalies.append(
                    f"**{q['name']}**: bot command has no summons_type:/badge_name: "
                    f"parameter — the bot cannot tell what is being turned in: `{cmd}`"
                )
            else:
                if credited in seen_commands and seen_commands[credited] != q["name"]:
                    anomalies.append(
                        f"**{q['name']}**: bot command credits `{credited}` which is "
                        f"also credited by **{seen_commands[credited]}** — likely a "
                        f"copy-paste slip in the source. The bot would credit the "
                        f"wrong quest."
                    )
                else:
                    seen_commands[credited] = q["name"]
                if credited.lower() != q["name"].lower():
                    anomalies.append(
                        f"**{q['name']}**: bot command credits `{credited}` — not the "
                        f"quest's own name. Typo or intentional?"
                    )
        cross_check(q, anomalies)

    return {
        "schema_version": 1,
        "guild": source["guild"],
        "era": source["era"],
        "source": {
            "kind": source["adapter"],
            "detail": f"{source.get('path')} :: {source.get('tab', '')}".strip(" :"),
            "url": source.get("url"),
            "retrieved": source.get("retrieved"),
        },
        "quests": quests,
    }


# ---------------------------------------------------------------- adapters

def adapt_sheet_xlsx(source):
    """Harvest a normalized quest tab from a guild tracker workbook.
    Expects columns: Name | Coopable? | Category | Turn-in Requirements | Bot Template."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("The sheet-xlsx adapter needs openpyxl: pip install openpyxl")

    path = os.path.normpath(os.path.join(HERE, source["path"]))
    wb = load_workbook(path, read_only=True)
    if source["tab"] not in wb.sheetnames:
        raise SystemExit(f"tab {source['tab']!r} not found in {path}")
    ws = wb[source["tab"]]

    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip().lower() if c else "" for c in rows[0]]
    expected = ["name", "coopable?", "category", "turn-in requirements", "bot template"]
    anomalies = []
    if header[: len(expected)] != expected:
        anomalies.append(
            f"header row is {header[:5]} — expected {expected}. "
            f"Columns may have moved; harvest read them positionally."
        )

    quests = []
    skipped = 0
    for i, row in enumerate(rows[1:], start=2):
        name = (str(row[0]).strip() if row[0] is not None else "")
        if not name:
            continue
        coopable, category, req, cmd = row[1], row[2], row[3], row[4]
        if category is None and req is None and cmd is None:
            anomalies.append(f"row {i}: has a name ({name!r}) but no other data — skipped.")
            skipped += 1
            continue
        req = str(req).strip() if req is not None else ""
        cmd = str(cmd).strip() if cmd is not None else None
        # the sheet marks meta-quests by putting "No submission, auto-checked" in the
        # template column instead of a command
        auto = (
            cmd is None
            or not cmd.startswith("/")
            or "auto-checked" in req.lower()
        )
        if cmd is not None and not cmd.startswith("/") and "auto" not in cmd.lower():
            anomalies.append(
                f"row {i} ({name!r}): bot template is not a command and does not say "
                f"auto-checked: {cmd!r}. Treated as auto-checked — is that right?"
            )
        quest = {
            "quest_id": slugify(name),
            "name": name,
            "category": str(category).strip() if category is not None else "",
            "coopable": bool(coopable),
            "requirements": req,
            "reward": None,
            "evidence": parse_evidence(None if auto else cmd, req),
            "evidence_note": None,
            "bot_command": None if auto else cmd,
            "auto_checked": auto,
            "venue": "in_game",
            "trigger": None,
        }
        quests.append(quest)

    if skipped:
        anomalies.append(f"{skipped} row(s) skipped for missing data (listed above).")
    return finish_catalog(source, quests, anomalies), anomalies


def adapt_gm_template(source):
    """A GM hands us a filled template that is already in catalog shape: validate the
    envelope, regenerate ids, run the same cross-checks. Content passes through."""
    path = os.path.normpath(os.path.join(HERE, source["path"]))
    with open(path, encoding="utf-8-sig") as f:
        data = json.load(f)
    anomalies = []
    quests = []
    for q in data.get("quests", []):
        name = (q.get("name") or "").strip()
        if not name:
            anomalies.append(f"a quest entry has no name — skipped: {json.dumps(q)[:80]}")
            continue
        cmd = q.get("bot_command")
        auto = bool(q.get("auto_checked")) or cmd is None
        quests.append({
            "quest_id": q.get("quest_id") or slugify(name),
            "name": name,
            "category": q.get("category", ""),
            "coopable": bool(q.get("coopable", False)),
            "requirements": q.get("requirements", ""),
            "evidence": q.get("evidence") or parse_evidence(None if auto else cmd, q.get("requirements", "")),
            "bot_command": None if auto else cmd,
            "auto_checked": auto,
            "venue": q.get("venue", "in_game"),
            "trigger": q.get("trigger"),
        })
    return finish_catalog(source, quests, anomalies), anomalies


def adapt_ranger_xlsx(source):
    """Harvest the Ranger tracker: a sectioned 'Badges' tab (category title rows, a
    'shared' group marker, per-badge evidence notes, an IRL section) plus a narrative
    'Quests' tab (name and reward folded into one cell)."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("The ranger-xlsx adapter needs openpyxl: pip install openpyxl")

    path = os.path.normpath(os.path.join(HERE, source["path"]))
    wb = load_workbook(path, read_only=True)
    anomalies = []
    quests = []

    # --- Badges tab: repeated sections, columns fixed at
    #     shared-marker=0, name=1, description=3, turn-in=7, evidence-note=10
    section = re.compile(r"^(.*\bBadges)\s*$")
    ws = wb["Badges"]
    category = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        cell0 = str(row[0]).strip() if row[0] is not None else ""
        name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        m = section.match(cell0)
        if m and not name:
            category = re.sub(r"\s+", " ", m.group(1))
            continue
        if not name or name == "Badge":
            continue
        # names sometimes carry a parenthesized annotation on a second line,
        # e.g. "Igloo\n(Ranger Station)" — the bot credits the bare name
        name_note = None
        name_lines = [ln.strip() for ln in name.splitlines() if ln.strip()]
        if len(name_lines) > 1 and all(ln.startswith("(") for ln in name_lines[1:]):
            name = name_lines[0]
            name_note = " ".join(name_lines[1:])
        if category is None:
            anomalies.append(f"Badges row {i} ({name!r}): appears before any section title — skipped.")
            continue
        desc = str(row[3]).strip() if len(row) > 3 and row[3] is not None else ""
        cmd = str(row[7]).strip() if len(row) > 7 and row[7] is not None else None
        note = str(row[10]).strip() if len(row) > 10 and row[10] is not None else None
        if not cmd:
            anomalies.append(f"Badges row {i} ({name!r}): no turn-in command — skipped. Auto-checked, retired, or a slip?")
            continue
        quests.append({
            "quest_id": slugify(name),
            "name": name,
            "name_note": name_note,
            "category": category,
            "coopable": "shared" in cell0.lower(),
            "requirements": desc,
            "reward": None,
            "evidence": parse_evidence(cmd, f"{desc}\n{note or ''}"),
            "evidence_note": note,
            "bot_command": cmd,
            "auto_checked": False,
            "venue": "irl" if category.lower().startswith("irl") else "in_game",
            "trigger": None,
        })

    # --- Quests tab: name + reward folded into col 0, description col 1, turn-in col 3
    ws = wb["Quests"]
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        blob = str(row[0]).strip() if row[0] is not None else ""
        cmd = str(row[3]).strip() if len(row) > 3 and row[3] is not None else None
        if not blob or blob == "Quest" or not cmd:
            continue
        lines = [ln.strip() for ln in blob.splitlines() if ln.strip()]
        name = lines[0] if lines else ""
        reward = None
        for j, ln in enumerate(lines):
            if ln.lower().startswith("reward"):
                inline = ln.split(":", 1)[1].strip() if ":" in ln else ""
                reward = inline or " ".join(lines[j + 1:]) or None
                break
        if not name:
            anomalies.append(f"Quests row {i}: has a turn-in command but no readable name — skipped: {blob[:60]!r}")
            continue
        desc = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
        quests.append({
            "quest_id": slugify(name),
            "name": name,
            "category": "Quests",
            "coopable": True,  # tab header: collaboration allowed and encouraged
            "requirements": desc,
            "reward": reward,
            "evidence": parse_evidence(cmd, desc),
            "evidence_note": None,
            "bot_command": cmd,
            "auto_checked": False,
            "venue": "in_game",
            "trigger": None,
        })

    return finish_catalog(source, quests, anomalies), anomalies


def adapt_discord_export(source):
    """Reserved: harvest quests straight from a Discord channel export. This is where
    the absorption engine plugs in. Deliberately unimplemented."""
    raise SystemExit(
        "The discord-export adapter is a stub — it marks the seam where the absorption "
        "engine will plug in. Use sheet-xlsx or gm-template for now."
    )


ADAPTERS = {
    "sheet-xlsx": adapt_sheet_xlsx,
    "ranger-xlsx": adapt_ranger_xlsx,
    "gm-template": adapt_gm_template,
    "discord-export": adapt_discord_export,
}


# ---------------------------------------------------------------- driver

def write_anomalies(path, source, anomalies):
    lines = [
        f"# Anomalies — {source['guild']} quest catalog ({source['id']})",
        "",
        "The harvester copies the guild's content verbatim and flags what looks off.",
        "Nothing here was 'fixed' — these are questions for the guild to rule on.",
        "",
    ]
    if anomalies:
        lines += [f"{i}. {a}" for i, a in enumerate(anomalies, 1)]
    else:
        lines.append("No anomalies found.")
    lines.append("")
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines))


def harvest(source):
    adapter = ADAPTERS.get(source["adapter"])
    if adapter is None:
        raise SystemExit(f"unknown adapter: {source['adapter']} (have: {', '.join(ADAPTERS)})")
    catalog, anomalies = adapter(source)

    out = os.path.normpath(os.path.join(HERE, source["output"]))
    os.makedirs(os.path.dirname(out), exist_ok=True)
    with open(out, "w", encoding="utf-8", newline="\n") as f:
        json.dump(catalog, f, indent=2, ensure_ascii=False)
        f.write("\n")

    anomalies_path = out[: -len(".json")] + "-anomalies.md" if out.endswith(".json") else out + "-anomalies.md"
    write_anomalies(anomalies_path, source, anomalies)

    print(f"[{source['id']}] {len(catalog['quests'])} quest(s) -> {os.path.relpath(out, HERE)}")
    print(f"[{source['id']}] {len(anomalies)} anomaly(ies) -> {os.path.relpath(anomalies_path, HERE)}")
    return catalog


def main():
    with open(os.path.join(HERE, "sources.json"), encoding="utf-8-sig") as f:
        config = json.load(f)

    wanted = sys.argv[1] if len(sys.argv) > 1 else None
    ran = 0
    for source in config["sources"]:
        if wanted is not None and source["id"] != wanted:
            continue
        if wanted is None and not source.get("enabled", True):
            print(f"[{source['id']}] disabled — skipped ({source.get('note', '')})")
            continue
        harvest(source)
        ran += 1

    if ran == 0:
        raise SystemExit(f"no source matched {wanted!r} — check sources.json")


if __name__ == "__main__":
    main()
